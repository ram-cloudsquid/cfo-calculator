#!/usr/bin/env python3
"""
Flask backend for CFO ROI Calculator - FIXED VERSION
Generates customized Excel workbooks based on user inputs from the HTML form

FIXES APPLIED:
1. Formulas now use Excel cell references (B4, B5, etc) instead of hardcoded Python values
2. This allows users to edit inputs and see formulas update automatically
3. Added scenario columns (Bear/Base/Bull) with proper formula references
"""

from flask import Flask, request, send_file, jsonify
import io
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

app = Flask(__name__)


@app.after_request
def add_cors_headers(response):
    """
    Manually set CORS headers instead of relying on the flask-cors package.
    This avoids failures when Flask-CORS isn't installed or a pinned version
    (e.g. 4.0.0) doesn't match the local Flask/Python setup — a common cause
    of 'blocked by CORS policy' errors even when the server is running fine.
    """
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response


@app.route('/api/generate-roi-report', methods=['OPTIONS'])
def generate_roi_report_preflight():
    """Explicitly answer the browser's CORS preflight OPTIONS request."""
    return '', 204


class ROIWorkbookGenerator:
    def __init__(self, data):
        self.mode = 'simple' if str(data.get('mode', 'advanced')).lower() == 'simple' else 'advanced'
        self.data = self._normalize_inputs(data)
        self.company_name = data.get('company_name', 'Company')
        self.revenue_m = float(data.get('revenue_m', 0))
        self.company_size = int(data.get('company_size', 100))
        self.industry = data.get('industry', 'general')
        self.region = data.get('region', 'us')
        self.verticals = data.get('verticals', [])
        self.created_date = datetime.now().strftime("%B %d, %Y")
        
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Color scheme
        self.COLOR_HEADER = "1F4E78"
        self.COLOR_SECTION = "D9E1F2"
        self.COLOR_INPUT = "FFF2CC"
        self.COLOR_CALC = "E2EFDA"
        self.COLOR_RESULT = "FCE4D6"
        self.COLOR_BENCH = "EDEDED"
        
        # Fonts
        self.font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        self.font_section = Font(name="Calibri", size=11, bold=True, color=self.COLOR_HEADER)
        self.font_subsection = Font(name="Calibri", size=10, bold=True, color="000000")
        self.font_input = Font(name="Calibri", size=11, color="0000FF")
        self.font_calc = Font(name="Calibri", size=11, color="000000")
        self.font_result = Font(name="Calibri", size=11, bold=True, color="C65911")
        self.font_label = Font(name="Calibri", size=10, color="595959")
        self.font_footnote = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
        
        # Fills
        self.fill_header = PatternFill(fill_type="solid", fgColor=self.COLOR_HEADER)
        self.fill_section = PatternFill(fill_type="solid", fgColor=self.COLOR_SECTION)
        self.fill_input = PatternFill(fill_type="solid", fgColor=self.COLOR_INPUT)
        self.fill_calc = PatternFill(fill_type="solid", fgColor=self.COLOR_CALC)
        self.fill_result = PatternFill(fill_type="solid", fgColor=self.COLOR_RESULT)
        self.fill_bench = PatternFill(fill_type="solid", fgColor=self.COLOR_BENCH)
        
        # Borders
        self.border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        self.border_thick = Border(
            left=Side(style='medium', color='1F4E78'),
            right=Side(style='medium', color='1F4E78'),
            top=Side(style='medium', color='1F4E78'),
            bottom=Side(style='medium', color='1F4E78')
        )
    
    # Defaults mirror the values already hardcoded into the HTML calculator's
    # <input value="..."> attributes, so a field left blank behaves the same
    # in the Excel as it does in the on-screen live preview.
    _ADVANCED_DEFAULTS = {
        'ap_volume': 20,
        'ap_3way_failure': 2.5, 'ap_3way_amount': 2500, 'ap_3way_recovery': 65,
        'ap_dup_rate': 0.4, 'ap_dup_amount': 3500, 'ap_dup_recovery': 90,
        'ap_fraud_rate': 0.3, 'ap_fraud_amount': 5000, 'ap_fraud_recovery': 45,
        'ap_credit_monthly': 15, 'ap_credit_not_applied': 12, 'ap_credit_recovery': 85,
        'ap_discount_rate': 1.5, 'ap_discount_eligible': 35, 'ap_discount_capture': 40,
        'freight_volume': 5,
        'freight_avg_shipment_value': 500,  # hardcoded in the live calculator's JS; exposed as an editable cell here
        'freight_auditable_pct': 65,
        'freight_misclass_rate': 18, 'freight_misclass_avg': 45,
        'freight_access_pct': 22, 'freight_access_invalid': 12, 'freight_access_avg': 38,
        'freight_rate_compliance': 68, 'freight_rate_overcharge': 8, 'freight_rate_recovery': 70,
        'freight_claim_rate': 1.2, 'freight_claim_avg': 1200, 'freight_claim_approval': 55,
        'rec_promo_rate': 0.5, 'rec_promo_valid': 70, 'rec_promo_recovery': 20,
        'rec_quality_rate': 0.3, 'rec_quality_valid': 55, 'rec_quality_recovery': 30,
        'rec_short_rate': 0.25, 'rec_short_valid': 85, 'rec_short_recovery': 45,
        'rec_pricing_rate': 0.15, 'rec_pricing_valid': 50, 'rec_pricing_recovery': 15,
        'rec_return_rate': 0.3, 'rec_return_recovery': 75,
        'rec_target_capture': 80, 'rec_prevention': 15,
    }

    _SIMPLE_DEFAULTS = {
        'ap_volume': 20, 'ap_problem_ratio': 3, 'ap_issue_amount': 2500, 'ap_recovery_rate': 70,
        'freight_volume': 5, 'freight_audit_coverage': 60, 'freight_overcharge_rate': 8, 'freight_recovery_rate': 75,
        'rec_deduction_rate': 1.5, 'rec_current_recovery': 35, 'rec_target_capture': 60, 'rec_prevention': 12,
    }

    # Bear/Bull multipliers, per vertical — these match the live on-screen
    # calculator's calculate*Simple/Advanced() functions exactly (same in
    # both modes), so the Excel scenarios agree with what the person saw
    # in the browser before downloading.
    SCENARIO_MULTIPLIERS = {
        'ap': (0.70, 1.35),
        'freight': (0.65, 1.40),
        'rec': (0.65, 1.55),
    }

    def _normalize_inputs(self, data):
        """Fill missing fields with the defaults for whichever mode (simple/advanced) was submitted."""
        normalized = dict(data)
        defaults = self._SIMPLE_DEFAULTS if self.mode == 'simple' else self._ADVANCED_DEFAULTS
        for key, default_value in defaults.items():
            if key not in normalized or normalized[key] in (None, ''):
                normalized[key] = default_value
        return normalized
    
    def _auto_adjust_columns(self, ws, start_row=1):
        """Auto-adjust column widths intelligently"""
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Track which columns have content
        columns_with_content = set()
        for row in ws.iter_rows(min_row=start_row, max_row=max_row):
            for idx, cell in enumerate(row, 1):
                if cell.value is not None and str(cell.value).strip():
                    columns_with_content.add(idx)
        
        # Adjust widths column by column
        for col_num in range(1, max_col + 1):
            col_letter = chr(64 + col_num)
            max_length = 0
            
            # Find max content length in this column
            for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=col_num, max_col=col_num):
                if row[0].value:
                    try:
                        cell_length = len(str(row[0].value))
                        max_length = max(max_length, cell_length)
                    except:
                        pass
            
            # Check if next column has content
            next_col_has_content = (col_num + 1) in columns_with_content
            
            # Set width based on content and next column status
            if max_length > 0:
                if next_col_has_content:
                    adjusted_width = min(max_length + 2, 55)
                else:
                    adjusted_width = min(max_length + 2, 25)
            else:
                adjusted_width = 12
            
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def _cell(self, ws, row, col, value, font=None, fill=None, num_format=None, 
              alignment=None, border=None, merge_cells=None):
        """Helper to set cell value with styling"""
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if num_format:
            cell.number_format = num_format
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if merge_cells:
            ws.merge_cells(merge_cells)
        return cell
    
    def add_cover_and_instructions(self):
        """Add cover page and instructions tab"""
        ws = self.wb.create_sheet("Welcome", 0)
        
        row = 1
        ws.row_dimensions[row].height = 36
        self._cell(ws, row, 1, "CFO ROI CALCULATOR", 
                  font=Font(name="Calibri", size=24, bold=True, color="FFFFFF"),
                  fill=self.fill_header, border=self.border_thick)
        
        row += 2
        self._cell(ws, row, 1, f"Prepared for: {self.company_name}", 
                  font=Font(name="Calibri", size=12, bold=True))
        row += 1
        self._cell(ws, row, 1, f"Date: {self.created_date}", 
                  font=Font(name="Calibri", size=11))
        
        row += 3
        self._cell(ws, row, 1, "Model Guide", font=self.font_section, fill=self.fill_section)
        row += 1
        
        instructions = [
            "This workbook models your potential financial recovery opportunities across three finance processes:",
            "",
            "1. ACCOUNTS PAYABLE – Identifying payment errors, duplicates, fraud, and missed early-pay discounts",
            "2. FREIGHT AUDIT – Finding shipment overcharges, misclassifications, and lost claims",
            "3. ACCOUNTS RECEIVABLE – Recovering invalid deductions, pricing disputes, and short shipments",
            "",
            "HOW TO USE THIS MODEL:",
            "• Company Inputs tab: Enter your business parameters (revenue, employee count, etc.)",
            "• Executive Summary tab: Review financial scenarios and recovery potential",
            "• Process tabs: Dive into detailed calculations for AP, Freight, or AR",
            "• All cells with yellow background are inputs – you can edit them",
            "• Formulas will auto-recalculate when inputs change",
            "",
            "SCENARIOS:",
            "• BEAR: Conservative estimate (30-35% of base)",
            "• BASE: Most likely outcome",
            "• BULL: Optimistic outcome (35-55% of base)",
        ]
        
        for text in instructions:
            self._cell(ws, row, 1, text, font=self.font_label, 
                      alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))
            ws.row_dimensions[row].height = 18
            row += 1
        
        self._auto_adjust_columns(ws)
        ws.column_dimensions['A'].width = 120
    
    def add_company_inputs(self):
        """Add company inputs tab"""
        ws = self.wb.create_sheet("Company Inputs", 1)
        
        row = 1
        ws.row_dimensions[row].height = 24
        self._cell(ws, row, 1, "COMPANY INFORMATION", 
                  font=self.font_header, fill=self.fill_header, border=self.border_thick)
        
        row += 2
        self._cell(ws, row, 1, "Company Name", font=self.font_subsection)
        self._cell(ws, row, 2, self.company_name, font=self.font_input, fill=self.fill_input)
        row += 1
        
        self._cell(ws, row, 1, "Annual Revenue ($M)", font=self.font_subsection)
        self._cell(ws, row, 2, self.revenue_m, font=self.font_input, fill=self.fill_input, num_format='$#,##0.0')
        row += 1
        
        self._cell(ws, row, 1, "Company Size (Employees)", font=self.font_subsection)
        self._cell(ws, row, 2, self.company_size, font=self.font_input, fill=self.fill_input, num_format='#,##0')
        row += 1
        
        self._cell(ws, row, 1, "Industry", font=self.font_subsection)
        self._cell(ws, row, 2, self.industry.title(), font=self.font_input, fill=self.fill_input)
        row += 1
        
        self._cell(ws, row, 1, "Region", font=self.font_subsection)
        self._cell(ws, row, 2, self.region.title(), font=self.font_input, fill=self.fill_input)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        self._auto_adjust_columns(ws)
    
    # Fixed row numbers where each detail tab writes its grand total.
    # These are deterministic per mode because each tab always has the same
    # fixed layout for that mode, regardless of the input data.
    _TOTAL_ROWS = {
        'advanced': {'ap': 35, 'freight': 32, 'rec': 46},
        'simple':   {'ap': 13, 'freight': 13, 'rec': 15},
    }

    def add_executive_summary(self):
        """Add executive summary tab: recovery-by-process, Bear/Base/Bull scenarios, key metrics."""
        ws = self.wb.create_sheet("Executive Summary", 2)
        total_rows = self._TOTAL_ROWS[self.mode]

        row = 1
        ws.row_dimensions[row].height = 24
        self._cell(ws, row, 1, "FINANCIAL RECOVERY SUMMARY",
                  font=self.font_header, fill=self.fill_header, border=self.border_thick)
        row += 1
        self._cell(ws, row, 1, f"Mode: {self.mode.title()}", font=self.font_footnote)

        # --- Recovery by process (cross-sheet links to each detail tab's total) ---
        row += 2
        self._cell(ws, row, 1, "Recovery by Process", font=self.font_section, fill=self.fill_section)
        row += 1
        self._cell(ws, row, 2, "Base $", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        self._cell(ws, row, 3, "Bear x", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        self._cell(ws, row, 4, "Bull x", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        row += 1

        has_rec = any(k in self.verticals for k in ('rec', 'ar', 'receivables'))
        # component: (label, sheet name, key into _TOTAL_ROWS/SCENARIO_MULTIPLIERS)
        components = []
        if 'ap' in self.verticals:
            components.append(("  Accounts Payable", "Accounts Payable", 'ap'))
        if 'freight' in self.verticals:
            components.append(("  Freight Audit", "Freight Audit", 'freight'))
        if has_rec:
            components.append(("  Receivables & Deductions", "Receivables & Deductions", 'rec'))

        base_rows, bear_mult_rows, bull_mult_rows = [], [], []
        for label, sheet_name, key in components:
            bear_default, bull_default = self.SCENARIO_MULTIPLIERS[key]
            self._cell(ws, row, 1, label)
            self._cell(ws, row, 2, f"='{sheet_name}'!B{total_rows[key]}",
                      font=self.font_calc, fill=self.fill_calc, num_format='$#,##0')
            base_rows.append(row)
            self._cell(ws, row, 3, bear_default, font=self.font_input, fill=self.fill_input, num_format='0.00"x"')
            bear_mult_rows.append(row)
            self._cell(ws, row, 4, bull_default, font=self.font_input, fill=self.fill_input, num_format='0.00"x"')
            bull_mult_rows.append(row)
            row += 1

        row += 1
        self._cell(ws, row, 1,
                  "Bear/Bull multipliers above match the live calculator's per-process defaults "
                  "(editable if you have a better basis for them).",
                  font=self.font_footnote, alignment=Alignment(wrap_text=True))
        ws.row_dimensions[row].height = 18
        row += 2

        # --- Total Recovery by Scenario ---
        self._cell(ws, row, 1, "Total Recovery by Scenario", font=self.font_section, fill=self.fill_section)
        row += 1
        self._cell(ws, row, 2, "Bear Case", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        self._cell(ws, row, 3, "Base Case", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        self._cell(ws, row, 4, "Bull Case", font=self.font_subsection, fill=self.fill_bench, alignment=Alignment(horizontal='center'))
        row += 1
        self._cell(ws, row, 1, "TOTAL ANNUAL RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
        if base_rows:
            bear_formula = "=" + "+".join(f"B{b}*C{b}" for b in base_rows)
            base_formula = "=" + "+".join(f"B{b}" for b in base_rows)
            bull_formula = "=" + "+".join(f"B{b}*D{b}" for b in base_rows)
        else:
            bear_formula = base_formula = bull_formula = 0
        self._cell(ws, row, 2, bear_formula, font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')
        self._cell(ws, row, 3, base_formula, font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')
        self._cell(ws, row, 4, bull_formula, font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')
        base_total_row = row
        row += 2

        # --- Key Metrics ---
        self._cell(ws, row, 1, "Key Metrics (Base Case)", font=self.font_section, fill=self.fill_section)
        row += 1
        self._cell(ws, row, 1, "  Recovery as % of Revenue")
        self._cell(ws, row, 2, f"=IFERROR(C{base_total_row}/('Company Inputs'!B4*1000000),0)",
                  font=self.font_result, fill=self.fill_calc, num_format='0.00%')
        row += 1
        self._cell(ws, row, 1, "  Recovery per Employee")
        self._cell(ws, row, 2, f"=IFERROR(C{base_total_row}/'Company Inputs'!B5,0)",
                  font=self.font_result, fill=self.fill_calc, num_format='$#,##0')
        row += 1
        self._cell(ws, row, 1, "  Payback Period (Months, 1 FTE @ $80K)")
        self._cell(ws, row, 2, f"=IFERROR(80000*12/C{base_total_row},0)",
                  font=self.font_result, fill=self.fill_calc, num_format='0.00')

        ws.column_dimensions['A'].width = 32
        self._auto_adjust_columns(ws)
    
    def _category_block(self, ws, row, title, fields, formula_fn):
        """
        Write one category block: a bold title, N input rows (label, value),
        a formula row, and return (row_after_block, result_row).

        fields: list of (label, data_key, default, num_format) tuples.
        formula_fn: function(row_map) -> formula string, where row_map maps
                    data_key -> the Excel row number that key was written to.
        """
        self._cell(ws, row, 1, title, font=Font(name="Calibri", size=11, bold=True, color="1F4E78"))
        row += 1
        row_map = {}
        for label, key, default, num_format in fields:
            self._cell(ws, row, 1, f"  {label}")
            self._cell(ws, row, 2, self.data.get(key, default),
                       font=self.font_input, fill=self.fill_input, num_format=num_format)
            row_map[key] = row
            row += 1
        self._cell(ws, row, 1, "  Recovery Potential", font=Font(name="Calibri", size=10, bold=True))
        self._cell(ws, row, 2, formula_fn(row_map),
                   font=self.font_result, fill=self.fill_result, num_format='$#,##0')
        result_row = row
        row += 2
        return row, result_row

    def _add_detail_tab_header(self, ws, title):
        row = 1
        ws.row_dimensions[row].height = 24
        self._cell(ws, row, 1, title, font=self.font_header, fill=self.fill_header, border=self.border_thick)
        row += 2
        return row

    def _add_ap_tab(self, tab_index):
        """Detailed Accounts Payable worksheet. Layout depends on mode."""
        ws = self.wb.create_sheet("Accounts Payable", tab_index)
        title = "ACCOUNTS PAYABLE — DETAIL" if self.mode == 'advanced' else "ACCOUNTS PAYABLE — SIMPLE ESTIMATE"
        row = self._add_detail_tab_header(ws, title)

        self._cell(ws, row, 1, "AP Volume ($M)")
        self._cell(ws, row, 2, self.data.get('ap_volume', 20), font=self.font_input, fill=self.fill_input, num_format='$#,##0')
        vol_row = row
        row += 2

        if self.mode == 'simple':
            self._cell(ws, row, 1, "Simple mode uses one combined issue rate rather than the 5-category "
                                    "breakdown below — switch to Advanced mode on the calculator for that level of detail.",
                       font=self.font_footnote, alignment=Alignment(wrap_text=True))
            ws.row_dimensions[row].height = 28
            row += 2

            row, r = self._category_block(ws, row, "OVERALL AP ISSUES", [
                ("Problem Ratio (%)", 'ap_problem_ratio', 3, '0.0'),
                ("Avg Issue Amount ($)", 'ap_issue_amount', 2500, '$#,##0'),
                ("Recovery Rate (%)", 'ap_recovery_rate', 70, '0.0'),
            ], lambda m: f"=(B{vol_row}*1000000/B{m['ap_issue_amount']}*B{m['ap_problem_ratio']}/100)*B{m['ap_issue_amount']}*B{m['ap_recovery_rate']}/100")

            self._cell(ws, row, 1, "TOTAL AP RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
            self._cell(ws, row, 2, f"=B{r}", font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

            ws.column_dimensions['A'].width = 32
            self._auto_adjust_columns(ws)
            return

        result_rows = []

        row, r = self._category_block(ws, row, "3-WAY MATCH FAILURES", [
            ("Failure Rate (%)", 'ap_3way_failure', 2.5, '0.0'),
            ("Avg Amount ($)", 'ap_3way_amount', 2500, '$#,##0'),
            ("Recovery Rate (%)", 'ap_3way_recovery', 65, '0.0'),
        ], lambda m: f"=B{vol_row}*1000000*(B{m['ap_3way_failure']}/100)*(B{m['ap_3way_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "DUPLICATE INVOICES", [
            ("Duplicate Rate (%)", 'ap_dup_rate', 0.4, '0.0'),
            ("Avg Amount ($)", 'ap_dup_amount', 3500, '$#,##0'),
            ("Recovery Rate (%)", 'ap_dup_recovery', 90, '0.0'),
        ], lambda m: f"=B{vol_row}*1000000*(B{m['ap_dup_rate']}/100)*(B{m['ap_dup_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "FRAUD", [
            ("Fraud Rate (%)", 'ap_fraud_rate', 0.3, '0.0'),
            ("Avg Amount ($)", 'ap_fraud_amount', 5000, '$#,##0'),
            ("Recovery Rate (%)", 'ap_fraud_recovery', 45, '0.0'),
        ], lambda m: f"=B{vol_row}*1000000*(B{m['ap_fraud_rate']}/100)*(B{m['ap_fraud_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "UNAPPLIED CREDITS", [
            ("Monthly Credits ($K)", 'ap_credit_monthly', 15, '$#,##0'),
            ("Not Applied (%)", 'ap_credit_not_applied', 12, '0.0'),
            ("Recovery Rate (%)", 'ap_credit_recovery', 85, '0.0'),
        ], lambda m: f"=B{m['ap_credit_monthly']}*1000*12*(B{m['ap_credit_not_applied']}/100)*(B{m['ap_credit_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "EARLY PAYMENT DISCOUNT CAPTURE", [
            ("Avg Discount (%)", 'ap_discount_rate', 1.5, '0.0'),
            ("% Invoices Eligible", 'ap_discount_eligible', 35, '0.0'),
            ("Current Capture Rate (%)", 'ap_discount_capture', 40, '0.0'),
        ], lambda m: f"=B{vol_row}*1000000*(B{m['ap_discount_eligible']}/100)*(B{m['ap_discount_rate']}/100)*(1-B{m['ap_discount_capture']}/100)")
        result_rows.append(r)

        self._cell(ws, row, 1, "TOTAL AP RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
        self._cell(ws, row, 2, "=" + "+".join(f"B{r}" for r in result_rows),
                   font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

        ws.column_dimensions['A'].width = 32
        self._auto_adjust_columns(ws)

    def _add_freight_tab(self, tab_index):
        """Detailed Freight Audit worksheet. Layout depends on mode."""
        ws = self.wb.create_sheet("Freight Audit", tab_index)
        title = "FREIGHT AUDIT — DETAIL" if self.mode == 'advanced' else "FREIGHT AUDIT — SIMPLE ESTIMATE"
        row = self._add_detail_tab_header(ws, title)

        self._cell(ws, row, 1, "Total Freight Spend ($M)")
        self._cell(ws, row, 2, self.data.get('freight_volume', 5), font=self.font_input, fill=self.fill_input, num_format='$#,##0')
        vol_row = row
        row += 2

        if self.mode == 'simple':
            self._cell(ws, row, 1, "Simple mode uses one combined audit-coverage/overcharge rate rather than the "
                                    "4-category breakdown below — switch to Advanced mode on the calculator for that level of detail.",
                       font=self.font_footnote, alignment=Alignment(wrap_text=True))
            ws.row_dimensions[row].height = 28
            row += 2

            row, r = self._category_block(ws, row, "OVERALL FREIGHT RECOVERY", [
                ("Audit Coverage (%)", 'freight_audit_coverage', 60, '0.0'),
                ("Overcharge Rate (%)", 'freight_overcharge_rate', 8, '0.0'),
                ("Recovery Rate (%)", 'freight_recovery_rate', 75, '0.0'),
            ], lambda m: f"=B{vol_row}*1000000*(B{m['freight_audit_coverage']}/100)*(B{m['freight_overcharge_rate']}/100)*(B{m['freight_recovery_rate']}/100)")

            self._cell(ws, row, 1, "TOTAL FREIGHT RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
            self._cell(ws, row, 2, f"=B{r}", font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

            ws.column_dimensions['A'].width = 32
            self._auto_adjust_columns(ws)
            return

        self._cell(ws, row, 1, "Avg Shipment Value ($)")
        self._cell(ws, row, 2, self.data.get('freight_avg_shipment_value', 500), font=self.font_input, fill=self.fill_input, num_format='$#,##0')
        avg_ship_row = row
        row += 1
        self._cell(ws, row, 1, "% of Spend Auditable")
        self._cell(ws, row, 2, self.data.get('freight_auditable_pct', 65), font=self.font_input, fill=self.fill_input, num_format='0.0')
        audit_row = row
        row += 1
        self._cell(ws, row, 1, "Auditable Shipments (count)", font=Font(name="Calibri", size=10, bold=True))
        self._cell(ws, row, 2, f"=B{vol_row}*1000000/B{avg_ship_row}*(B{audit_row}/100)", font=self.font_calc, fill=self.fill_calc, num_format='#,##0')
        audited_ships_row = row
        row += 2

        result_rows = []

        # Misclassification and Accessorial both draw from the auditable-shipment
        # pool; Rate Compliance and Claims apply against total spend/shipments
        # directly (no auditable-% filter) — matching the live calculator.
        row, r = self._category_block(ws, row, "WEIGHT & ZONE MISCLASSIFICATION", [
            ("Misclassification Rate (%)", 'freight_misclass_rate', 18, '0.0'),
            ("Avg Overcharge ($)", 'freight_misclass_avg', 45, '$#,##0'),
        ], lambda m: f"=B{audited_ships_row}*(B{m['freight_misclass_rate']}/100)*B{m['freight_misclass_avg']}")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "ACCESSORIAL CHARGES", [
            ("% Shipments w/ Accessorials", 'freight_access_pct', 22, '0.0'),
            ("Invalid Accessorial Rate (%)", 'freight_access_invalid', 12, '0.0'),
            ("Avg Amount ($)", 'freight_access_avg', 38, '$#,##0'),
        ], lambda m: f"=B{audited_ships_row}*(B{m['freight_access_pct']}/100)*(B{m['freight_access_invalid']}/100)*B{m['freight_access_avg']}")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "RATE NON-COMPLIANCE (vs Contract)", [
            ("Rate Compliance (%)", 'freight_rate_compliance', 68, '0.0'),
            ("Overcharge Rate (%)", 'freight_rate_overcharge', 8, '0.0'),
            ("Recovery Rate (%)", 'freight_rate_recovery', 70, '0.0'),
        ], lambda m: f"=B{vol_row}*1000000*(1-B{m['freight_rate_compliance']}/100)*(B{m['freight_rate_overcharge']}/100)*(B{m['freight_rate_recovery']}/100)")
        result_rows.append(r)

        self._cell(ws, row, 1, "LOSS & DAMAGE CLAIMS", font=Font(name="Calibri", size=11, bold=True, color="1F4E78"))
        row += 1
        self._cell(ws, row, 1, "  Claim Rate (%)")
        self._cell(ws, row, 2, self.data.get('freight_claim_rate', 1.2), font=self.font_input, fill=self.fill_input, num_format='0.0')
        claim_rate_row = row
        row += 1
        self._cell(ws, row, 1, "  Avg Claim ($)")
        self._cell(ws, row, 2, self.data.get('freight_claim_avg', 1200), font=self.font_input, fill=self.fill_input, num_format='$#,##0')
        claim_avg_row = row
        row += 1
        self._cell(ws, row, 1, "  Carrier Approval Rate (%)")
        self._cell(ws, row, 2, self.data.get('freight_claim_approval', 55), font=self.font_input, fill=self.fill_input, num_format='0.0')
        claim_approval_row = row
        row += 1
        self._cell(ws, row, 1, "  Recovery Potential", font=Font(name="Calibri", size=10, bold=True))
        # Claims apply to ALL shipments (spend / avg shipment value), not just the auditable subset.
        self._cell(ws, row, 2, f"=(B{vol_row}*1000000/B{avg_ship_row})*(B{claim_rate_row}/100)*B{claim_avg_row}*(B{claim_approval_row}/100)",
                   font=self.font_result, fill=self.fill_result, num_format='$#,##0')
        result_rows.append(row)
        row += 2

        self._cell(ws, row, 1, "TOTAL FREIGHT RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
        self._cell(ws, row, 2, "=" + "+".join(f"B{r}" for r in result_rows),
                   font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

        ws.column_dimensions['A'].width = 32
        self._auto_adjust_columns(ws)

    def _add_receivables_tab(self, tab_index):
        """Detailed Receivables & Deductions worksheet. Layout depends on mode."""
        ws = self.wb.create_sheet("Receivables & Deductions", tab_index)
        title = "RECEIVABLES & DEDUCTIONS — DETAIL" if self.mode == 'advanced' else "RECEIVABLES & DEDUCTIONS — SIMPLE ESTIMATE"
        row = self._add_detail_tab_header(ws, title)

        self._cell(ws, row, 1, "Annual Revenue ($M)")
        self._cell(ws, row, 2, self.revenue_m, font=self.font_input, fill=self.fill_input, num_format='$#,##0.0')
        rev_row = row
        row += 2

        if self.mode == 'simple':
            self._cell(ws, row, 1, "Simple mode uses one combined deduction rate rather than the 5-category "
                                    "breakdown below — switch to Advanced mode on the calculator for that level of detail.",
                       font=self.font_footnote, alignment=Alignment(wrap_text=True))
            ws.row_dimensions[row].height = 28
            row += 2

            self._cell(ws, row, 1, "  Deduction Rate (% of revenue)")
            self._cell(ws, row, 2, self.data.get('rec_deduction_rate', 1.5), font=self.font_input, fill=self.fill_input, num_format='0.00')
            ded_rate_row = row
            row += 1
            self._cell(ws, row, 1, "  Current Recovery Rate (%)")
            self._cell(ws, row, 2, self.data.get('rec_current_recovery', 35), font=self.font_input, fill=self.fill_input, num_format='0.0')
            cur_rec_row = row
            row += 1
            self._cell(ws, row, 1, "  Total Deductions ($)", font=Font(name="Calibri", size=10, bold=True))
            self._cell(ws, row, 2, f"=B{rev_row}*1000000*(B{ded_rate_row}/100)", font=self.font_calc, fill=self.fill_calc, num_format='$#,##0')
            ded_row = row
            row += 1
            self._cell(ws, row, 1, "  Unrecovered (Opportunity)", font=Font(name="Calibri", size=10, bold=True))
            self._cell(ws, row, 2, f"=B{ded_row}*(1-B{cur_rec_row}/100)", font=self.font_calc, fill=self.fill_calc, num_format='$#,##0')
            unrec_row = row
            row += 2

            self._cell(ws, row, 1, "  Target Capture Rate (%)")
            self._cell(ws, row, 2, self.data.get('rec_target_capture', 60), font=self.font_input, fill=self.fill_input, num_format='0.0')
            capture_row = row
            row += 1
            self._cell(ws, row, 1, "  Prevention Rate (% of deductions)")
            self._cell(ws, row, 2, self.data.get('rec_prevention', 12), font=self.font_input, fill=self.fill_input, num_format='0.0')
            prevent_row = row
            row += 2

            self._cell(ws, row, 1, "TOTAL RECEIVABLES RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
            self._cell(ws, row, 2, f"=B{unrec_row}*(B{capture_row}/100)+B{ded_row}*(B{prevent_row}/100)",
                       font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

            ws.column_dimensions['A'].width = 32
            self._auto_adjust_columns(ws)
            return

        self._cell(ws, row, 1, "Note: each category below is the currently-UNrecovered opportunity "
                                "(amount × valid% × (1 − current recovery%)). The Target Capture Rate "
                                "and Prevention Rate then convert that opportunity into projected recovery.",
                   font=self.font_footnote, alignment=Alignment(wrap_text=True))
        ws.row_dimensions[row].height = 28
        row += 2

        result_rows = []

        row, r = self._category_block(ws, row, "PROMOTIONAL / TRADE DEDUCTIONS", [
            ("Deduction Rate (%)", 'rec_promo_rate', 0.5, '0.00'),
            ("% Valid (Legitimate)", 'rec_promo_valid', 70, '0.0'),
            ("Current Recovery Rate (%)", 'rec_promo_recovery', 20, '0.0'),
        ], lambda m: f"=B{rev_row}*1000000*(B{m['rec_promo_rate']}/100)*(B{m['rec_promo_valid']}/100)*(1-B{m['rec_promo_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "QUALITY/DAMAGE DISPUTES", [
            ("Deduction Rate (%)", 'rec_quality_rate', 0.3, '0.00'),
            ("% Valid (Actually Damaged)", 'rec_quality_valid', 55, '0.0'),
            ("Current Recovery Rate (%)", 'rec_quality_recovery', 30, '0.0'),
        ], lambda m: f"=B{rev_row}*1000000*(B{m['rec_quality_rate']}/100)*(B{m['rec_quality_valid']}/100)*(1-B{m['rec_quality_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "SHORT QUANTITY / MISSING ITEMS", [
            ("Deduction Rate (%)", 'rec_short_rate', 0.25, '0.00'),
            ("% Valid (Actually Shipped Short)", 'rec_short_valid', 85, '0.0'),
            ("Current Recovery Rate (%)", 'rec_short_recovery', 45, '0.0'),
        ], lambda m: f"=B{rev_row}*1000000*(B{m['rec_short_rate']}/100)*(B{m['rec_short_valid']}/100)*(1-B{m['rec_short_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "PRICING / OVERCHARGE DISPUTES", [
            ("Deduction Rate (%)", 'rec_pricing_rate', 0.15, '0.00'),
            ("% Valid (We Did Overbill)", 'rec_pricing_valid', 50, '0.0'),
            ("Current Recovery Rate (%)", 'rec_pricing_recovery', 15, '0.0'),
        ], lambda m: f"=B{rev_row}*1000000*(B{m['rec_pricing_rate']}/100)*(B{m['rec_pricing_valid']}/100)*(1-B{m['rec_pricing_recovery']}/100)")
        result_rows.append(r)

        row, r = self._category_block(ws, row, "AUTHORIZED CUSTOMER RETURNS", [
            ("Return Rate (%)", 'rec_return_rate', 0.3, '0.00'),
            ("Current Recovery Rate (%)", 'rec_return_recovery', 75, '0.0'),
        ], lambda m: f"=B{rev_row}*1000000*(B{m['rec_return_rate']}/100)*(1-B{m['rec_return_recovery']}/100)")
        result_rows.append(r)

        self._cell(ws, row, 1, "TOTAL UNRECOVERED DEDUCTIONS", font=Font(name="Calibri", size=11, bold=True), fill=self.fill_bench)
        self._cell(ws, row, 2, "=" + "+".join(f"B{r}" for r in result_rows),
                   font=Font(name="Calibri", size=11, bold=True), fill=self.fill_calc, num_format='$#,##0')
        unrec_total_row = row
        row += 2

        self._cell(ws, row, 1, "INCREMENTAL RECOVERY (Target Capture)", font=Font(name="Calibri", size=11, bold=True, color="1F4E78"))
        row += 1
        self._cell(ws, row, 1, "  Target Capture Rate (%)")
        self._cell(ws, row, 2, self.data.get('rec_target_capture', 80), font=self.font_input, fill=self.fill_input, num_format='0.0')
        capture_row = row
        row += 1
        self._cell(ws, row, 1, "  Incremental Recovery from Capture", font=Font(name="Calibri", size=10, bold=True))
        self._cell(ws, row, 2, f"=B{unrec_total_row}*(B{capture_row}/100)", font=self.font_result, fill=self.fill_result, num_format='$#,##0')
        incremental_row = row
        row += 2

        self._cell(ws, row, 1, "DEDUCTION PREVENTION (Better Data/Processes)", font=Font(name="Calibri", size=11, bold=True, color="1F4E78"))
        row += 1
        self._cell(ws, row, 1, "  Prevention Rate (% of deductions avoided)")
        self._cell(ws, row, 2, self.data.get('rec_prevention', 15), font=self.font_input, fill=self.fill_input, num_format='0.0')
        prevent_row = row
        row += 1
        self._cell(ws, row, 1, "  Prevention Amount", font=Font(name="Calibri", size=10, bold=True))
        self._cell(ws, row, 2, f"=B{unrec_total_row}*(B{prevent_row}/100)", font=self.font_result, fill=self.fill_result, num_format='$#,##0')
        prevention_amount_row = row
        row += 2

        self._cell(ws, row, 1, "TOTAL RECEIVABLES RECOVERY", font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"), fill=self.fill_header)
        self._cell(ws, row, 2, f"=B{incremental_row}+B{prevention_amount_row}",
                   font=Font(name="Calibri", size=11, bold=True), fill=self.fill_result, num_format='$#,##0')

        ws.column_dimensions['A'].width = 32
        self._auto_adjust_columns(ws)

    def add_process_tabs(self):
        """Add detailed process tabs for each selected vertical."""
        tab_index = 3

        if 'ap' in self.verticals:
            self._add_ap_tab(tab_index)
            tab_index += 1

        if 'freight' in self.verticals:
            self._add_freight_tab(tab_index)
            tab_index += 1

        if 'rec' in self.verticals or 'ar' in self.verticals or 'receivables' in self.verticals:
            self._add_receivables_tab(tab_index)
            tab_index += 1
    
    def add_defaults_methodology(self):
        """Add defaults and methodology tab"""
        ws = self.wb.create_sheet("Benchmarks & Methodology", 7)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        
        row = 1
        ws.row_dimensions[row].height = 24
        self._cell(ws, row, 1, "BENCHMARKS & METHODOLOGY", 
                  font=self.font_header, fill=self.fill_header, border=self.border_thick)
        
        row += 2
        self._cell(ws, row, 1, "cloudsquid Methodology", font=self.font_section, fill=self.fill_section)
        row += 1
        methodology = [
            "These benchmarks were derived from:",
            "• Industry reports (NACM, IMA, NAAC)",
            "• Freight audit data from major carriers (FedEx, UPS, XPO, Old Dominion)",
            "• Payment processing data (SAP, Oracle, NetSuite implementations)",
            "",
            "Ranges shown reflect 25th-75th percentile across all observations.",
            "Your actual rates may vary based on company size, industry, and process maturity.",
        ]
        
        for text in methodology:
            self._cell(ws, row, 1, text, font=self.font_label, alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))
            ws.row_dimensions[row].height = 18
            row += 1
        
        self._auto_adjust_columns(ws)
        ws.column_dimensions['A'].width = 120

    def add_next_steps(self):
        """Add implementation roadmap tab"""
        ws = self.wb.create_sheet("Next Steps", 8)
        ws.column_dimensions['A'].width = 50
        
        row = 1
        ws.row_dimensions[row].height = 24
        self._cell(ws, row, 1, "IMPLEMENTATION ROADMAP & NEXT STEPS", 
                  font=self.font_header, fill=self.fill_header, border=self.border_thick)
        
        row += 2
        self._cell(ws, row, 1, "Recommended 90-Day Plan", font=self.font_section, fill=self.fill_section)
        row += 1
        
        phases = [
            ("Phase 1: Quick Win (Weeks 1–2)", "Data audit & readiness assessment", 
             "Inventory current processes, identify data gaps, assess team capacity"),
            ("Phase 2: Pilot (Weeks 3–6)", "Test AI agent on limited volume", 
             "Process sample invoices/shipments through exception detection engine"),
            ("Phase 3: Scale (Weeks 7–12)", "Deploy to full transaction population", 
             "Integrate with AP/freight/AR systems, train team, set up workflows"),
            ("Phase 4: Optimize (Weeks 13+)", "Refine rules & capture incremental gains", 
             "Iterate on thresholds, finalize recovery claims, measure results"),
        ]
        
        for phase, activity, detail in phases:
            self._cell(ws, row, 1, phase, font=self.font_subsection)
            ws.row_dimensions[row].height = 18
            row += 1
            self._cell(ws, row, 1, detail, font=self.font_footnote,
                      alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))
            ws.row_dimensions[row].height = 18
            row += 1
        
        row += 1
        self._cell(ws, row, 1, "Success Factors", font=self.font_section, fill=self.fill_section)
        row += 1
        
        factors = [
            "✓ Executive sponsor alignment – commitment from CFO/VP Finance",
            "✓ Clean data – invoices/transactions must be accurate for AI to work",
            "✓ Process definition – clear exception handling and approval workflows",
            "✓ Team training – prepare AP/AR staff for new tools and productivity gains",
            "✓ Weekly measurement – track recovered $ to build momentum",
        ]
        
        for factor in factors:
            self._cell(ws, row, 1, factor, font=self.font_label)
            ws.row_dimensions[row].height = 18
            row += 1
        
        row += 1
        self._cell(ws, row, 1, "Contact cloudsquid", font=self.font_section, fill=self.fill_section)
        row += 1
        
        self._cell(ws, row, 1, "Email: hello@cloudsquid.io")
        row += 1
        self._cell(ws, row, 1, "Website: https://cloudsquid.io")
        row += 1
        self._cell(ws, row, 1, "Calendar: https://calendly.com/cloudsquid/lookback")
        
        self._auto_adjust_columns(ws)
        ws.column_dimensions['A'].width = 120
    
    def generate(self):
        """Generate the complete workbook"""
        self.add_cover_and_instructions()
        self.add_company_inputs()
        self.add_executive_summary()
        self.add_process_tabs()
        self.add_defaults_methodology()
        self.add_next_steps()
        
        # Return as bytes
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()


@app.route('/api/generate-roi-report', methods=['POST'])
def generate_roi_report():
    """Generate and download ROI report"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required = ['company_name', 'revenue_m', 'company_size', 'industry', 'region', 'verticals']
        for field in required:
            if field not in data:
                return jsonify({'error': f'Missing field: {field}'}), 400
        
        # Generate workbook
        generator = ROIWorkbookGenerator(data)
        workbook_bytes = generator.generate()
        
        # Return as downloadable file
        return send_file(
            io.BytesIO(workbook_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"CFO_ROI_Model_{generator.mode.title()}_{data['company_name'].replace(' ', '_')}.xlsx"
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'ok'}), 200


if __name__ == '__main__':
    app.run(debug=True, port=5001)
